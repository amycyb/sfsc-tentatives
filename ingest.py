#!/usr/bin/env python3
"""
Ingest tentative rulings and maintain tentatives.parquet + tentatives.db.

Usage:
    python ingest.py                              # rebuild from all source files
    python ingest.py path/to/new.xlsx             # append a new export
    python ingest.py --dept 525 path/to/525.xlsx  # tag the rows with a department
    python ingest.py path/to/new.json             # append a new json export (extension output)

tentatives.parquet  — canonical dataset, committed to git (~10 MB)
tentatives.db       — local SQLite for querying, gitignored (~100 MB)

JSON formats accepted:
  Legacy array:  [{Case Number, Case Title, Court Date, Calendar Matter, Rulings}, ...]
                 (dept must be passed via --dept; defaults to 302 for back-compat)
  Extension:     {department, scraped_at, source_url, rulings: [...]}
                 (dept comes from the wrapper; --dept is ignored if present)
"""

import sys
import json
import re
import argparse
import sqlite3
import hashlib
from pathlib import Path
from datetime import datetime, date, time

try:
    import pandas as pd
except ImportError:
    sys.exit("Missing dependency: pip install pandas pyarrow openpyxl")

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl")

HERE    = Path(__file__).parent
PARQUET = HERE / "tentatives.parquet"
DB_PATH = HERE / "tentatives.db"

# Load judge code → name mapping from judges.json
_judges_path = HERE / "judges.json"
if _judges_path.exists():
    _judge_db = json.loads(_judges_path.read_text())
    JUDGE_CODE_MAP = {k: v["name"] for k, v in _judge_db.get("code_map", {}).items()}
else:
    JUDGE_CODE_MAP = {}


_JUDGE_SUFFIX_COMMA_RE = re.compile(r',\s*((?:Jr|Sr|II|III|IV)\.?)', re.IGNORECASE)
_WS_RE = re.compile(r'\s+')


def normalize_judge_name(name):
    """Canonicalise judge names so Excel and extension-scrape variants merge.

    Strips the optional comma before generational suffixes ("Richard B. Ulmer,
    Jr." → "Richard B. Ulmer Jr.") and collapses runs of whitespace. Without
    this the dropdown shows two Ulmer rows: Excel imports use the comma
    form, JSON scrapes use the no-comma form (extension JUDGE_MAP convention),
    and they live as distinct strings in the parquet.
    """
    if not name or not isinstance(name, str):
        return name
    stripped = _JUDGE_SUFFIX_COMMA_RE.sub(r' \1', name)
    return _WS_RE.sub(' ', stripped).strip()


def extract_judge(ruling_text):
    if not ruling_text:
        return None
    # Trailing tag forms observed in the wild:
    #   =(302/CK)  =(D302/CK)  (302/CK)  =(JPT)  =(525/JPT)  =(JPT/525)
    #   +(302/HEK) =(302.JMQ)  =(HEK)    (D302)  =(D525)     =(525)
    #
    # We require *either* an `=`/`+` prefix on the parenthetical *or* a
    # dept-vs-code separator (/, .) inside the parens. Otherwise an
    # ordinary trailing reference like "(CCP 1094.5)" — Code-of-Civil-
    # Procedure citations, common in petition rulings — gets read as a
    # fictitious "CCP" judge and the row's `judge` ends up null instead
    # of being populated by other heuristics.
    m = re.search(r'([=+])?\s*\(\s*([A-Za-z0-9][A-Za-z0-9\s/.,]{0,15})\s*\)\s*\.?\s*$', ruling_text)
    if not m:
        return None
    has_prefix = m.group(1) is not None
    inside = m.group(2)
    has_sep = bool(re.search(r'[\/.]', inside))
    if not has_prefix and not has_sep:
        return None
    # Pick the first letter-only run that isn't a bare D dept-marker.
    # Dept-only tags like (D302) yield no code → None (data genuinely lacks a judge code).
    code = next((c.upper() for c in re.findall(r'[A-Za-z]+', inside) if c.upper() != 'D'), None)
    if not code:
        return None
    if code == 'JPT':
        pt = re.search(
            r'Pro Tem Judge\s+([A-Z][A-Za-z.]+(?:\s+[A-Z][A-Za-z.]+)*?)'
            r'(?:,|;|\s+a\s+member|\s+member|\s+has been|\s+recuses)',
            ruling_text
        )
        if pt:
            return f"Judge Pro Tem: {pt.group(1).strip()}"
        return "Judge Pro Tem"
    return JUDGE_CODE_MAP.get(code)

COLUMNS = ["department", "case_number", "case_title", "court_date", "hearing_time",
           "calendar_matter", "judge", "ruling", "row_hash", "calendar_kind"]

# ─── Admin / CourtCall pattern set ──────────────────────────────────────────
# Keep in sync with the ADMIN_SENTENCE_PATTERNS + COURTCALL_RE in
# index.html. Pre-splitting at ingest time means the per-dept parquet
# can ship just the substantive ruling — admin / courtcall live in a
# sidecar file the data browser only fetches when a user opens a
# modal and expands the collapsible. That's how the dataset stays
# under wire-budget on first load.
_ADMIN_PATTERNS = [re.compile(p, re.IGNORECASE) for p in [
    # Contestation procedure
    r'\bany\s+part(?:y|ies)\s+who\s+contests?\s+a\s+tentative\s+ruling\b',
    r'\bif\s+(?:a|any|the)\s+part(?:y|ies)\s+(?:intends|wishes|wants|desires)\s+to\s+contest\b',
    r'\bto\s+contest\s+(?:the|this)\s+tentative\b',
    r'\bany\s+part(?:y|ies)\s+(?:intending|wishing)\s+to\s+(?:oppose|contest)\b',
    r'\btentative\s+ruling[s]?\s+will\s+become\s+(?:the\s+)?final\b',
    r'\bnotice\s+(?:of\s+)?contesting\s+a\s+tentative\b',
    r'\bnotice\s+of\s+intent(?:ion)?\s+to\s+(?:appear|contest)\b',
    r'\bwithout\s+argument,?\s+the\s+portion',
    # Procedural / appearance instructions
    r'\bparties\s+may\s+appear\s+(?:in[\s-]?person|telephonically|via\s+zoom|remotely)',
    r'\bparties\s+who\s+intend(?:\s+to)?\s+appear',
    r'\bparties\s+(?:must|shall)\s+give\s+notice\b',
    r'\bparties\s+who\s+intend\s+to\s+appear\s+at\s+the\s+hearing\b',
    r'\ba\s+party\s+may\s+not\s+argue\s+at\s+the\s+hearing\b',
    r'\bopposing\s+party\s+(?:does\s+not|is\s+not\s+so\s+notified)',
    r'\bface\s+coverings\s+are\s+(?:optional|required)',
    r'\bpursuant\s+to\s+(?:CRC|California\s+Rule(?:s)?\s+of\s+Court|Local\s+Rule)',
    r'\ball\s+attorneys?\s+(?:and\s+parties\s+)?may\s+appear\b',
    r'\b(?:may|shall)\s+appear\s+in\s+(?:department|dept\.?)\s*\d+\b',
    r'\b(?:8|9|10|11)\s*:\s*\d{2}\s*(?:a\.?m\.?|p\.?m\.?)\s+(?:law\s*(?:&|and)\s*motion|discovery|civil|calendar|department)\b',
    r'\bremote\s+hearings?\s+(?:will\s+be|are)\s+conducted\b',
    r'\bto\s+appear\s+remotely\s+at\s+the\s+hearing\b',
    r'\bgo\s+to\s+the\s+court\'?s?\s+website\b',
    r'\bdial\s+the\s+(?:corresponding|appropriate)\s+phone\s+number\b',
    r'\bnavigate\s+to\s+["\']?tentative\s+rulings',
    r'\bclick\s+on\s+the\s+(?:appropriate|corresponding)\s+(?:link|number)\b',
    r'\bemail\s+shall\s+include\s+the\b',
    r'\b(?:subject|text)\s+(?:line\s+)?of\s+the\s+email\b',
    r'\bsend\s+an?\s+email\s+to\b',
    r'\bcopy\s+to\s+all\s+(?:other\s+)?parties\s+by\s+\d',
    r'\bportion\(s\)\s+of\s+the\s+tentative\s+ruling\s+that\s+the\s+party\s+contests',
    r'\bof\s+the\s+attorney\s+or\s+party\s+who\s+will\s+appear\b',
    # Order-preparation boilerplate
    r'\bcounsel\s+for\s+the\s+prevailing\s+party\b.{0,40}\b(?:prepare|submit|propose)\b',
    r'\bprevailing\s+party\s+(?:is\s+(?:required|directed)|shall|must)\s+(?:to\s+)?(?:prepare|submit|lodge)\b',
    r'\bproposed\s+order\s+(?:repeating|that\s+repeats|repeats)\s+verbatim\b',
    r'\blodge\s+with\s+the\s+clerk\s+(?:in\s+)?(?:dept|department)\s*\d+\b',
    r'\bsubmit\s+a\s+proposed\s+order\b',
    # Court reporter info
    r'\bthe\s+court\s+(?:no\s+longer\s+)?provides?\s+a\s+court\s+reporter\b',
    r'\bparties\s+may\s+retain\s+(?:their\s+)?own\s+(?:court\s+)?reporter\b',
    r'\ba\s+retained\s+reporter\s+must\s+be\s+a\s+california\s+certified',
    r'\bif\s+a\s+csr\s+is\s+being\s+retained\b',
    r'\bcalifornia\s+certified\s+(?:shorthand\s+)?reporter\b',
    # Remote-appearance / Zoom / dial-in
    r'\bzoom\s+(?:meeting|webinar|video|id)',
    r'\bvideoconference\b',
    r'\bcourtcall\b',
    r'\bappear\s+(?:in[\s-]?person|telephonically|via\s+zoom|remotely)\b',
    r'\bwebinar\s+id\b',
    r'\bmeeting\s+id\b',
    r'\bphone\s+dial\s+in\b',
    r'\bdial[\s-]?in\s+(?:number|info|id|info\s+for)',
    r'\bsfsuperiorcourt\.org\b',
    r'\bmicrosoft\s+teams\b',
    r'\bwebex\b',
    # Email contest addresses
    r'\b[\w.+-]+@sftc\.org\b',
    # Time-based filing deadline boilerplate
    r'\bno\s+later\s+than\s+\d{1,2}\s*:\s*\d{2}\s*(?:a\.?m\.?|p\.?m\.?)\s+(?:the\s+)?(?:court\s+)?day\s+(?:before|prior\s+to)',
    r'\bby\s+\d{1,2}\s*:\s*\d{2}\s*(?:a\.?m\.?|p\.?m\.?)\s+(?:the\s+)?(?:court\s+)?day\s+(?:before|prior\s+to)',
    r'\bby\s+\d{1,2}\s*:\s*\d{2}\s*(?:a\.?m\.?|p\.?m\.?)\s+on\s+the\s+(?:court\s+)?day\s+(?:before|prior\s+to)',
]]

_COURTCALL_RE = re.compile(
    r"(?:CourtCall|(?:8|9|10|11)\s*:\s*\d{2}\s*a\.?m\.?\s+(?:law\s*(?:&|and)\s*motion|discovery|civil|calendar|department)"
    r"|appear\s+remotely|appear\s+(?:in[\s-]?person|telephonically|via\s+zoom)|\bzoom\b|videoconference|sfsuperiorcourt\.org"
    r"|telephonic\s+appearance|remote\s+appearance|microsoft\s+teams\b|webex\b|webinar\s+id\b|meeting\s+id\b|dial[\s-]?in\b"
    r"|phone\s+dial\s+in|all\s+attorneys?\s+(?:and\s+parties\s+)?may\s+appear|(?:may|shall)\s+appear\s+in\s+(?:department|dept\.?)\s*\d+"
    r"|remote\s+hearings?\s+(?:will\s+be|are)\s+conducted|to\s+appear\s+remotely\s+at\s+the\s+hearing|go\s+to\s+the\s+court'?s?\s+website"
    r"|dial\s+the\s+(?:corresponding|appropriate)\s+phone\s+number|click\s+on\s+the\s+(?:appropriate|corresponding)\s+(?:link|number))",
    re.IGNORECASE,
)

_SENTENCE_SPLIT_RE = re.compile(r'(?<=[.!?])\s+(?=[A-Z(=])')

def split_ruling(full):
    """Same shape as the JS splitRuling: returns (substantive, admin,
    courtcall) for a ruling text. Sentence-level classification — each
    sentence falls into exactly one bucket. Admin sentences that also
    match COURTCALL_RE get routed to the courtcall bucket so the modal
    can label them separately."""
    if not full:
        return ('', '', '')
    sentences = _SENTENCE_SPLIT_RE.split(full)
    sub_, adm_, cc_ = [], [], []
    for s in sentences:
        s = s.strip()
        if not s:
            continue
        if any(p.search(s) for p in _ADMIN_PATTERNS):
            if _COURTCALL_RE.search(s):
                cc_.append(s)
            else:
                adm_.append(s)
        else:
            sub_.append(s)
    return (
        re.sub(r'[\s.,;:]+$', '.', ' '.join(sub_)).strip(),
        ' '.join(adm_).strip(),
        ' '.join(cc_).strip(),
    )

# Suffixes that mark a calendar_matter as part of a split ruling. Each
# regex matches at the END of the calendar_matter so we strip them and
# get a clean canonical form to group the parts under. Many continuation
# rows have NO part numbers in their suffix (e.g. SFTC's
# "(ADDED TO CALENDAR FOR TENTATIVE RULING ENTRY PURPOSES ONLY.)" tag,
# or the bare "(continues, see next entry)" / "(continued from previous
# entry)" markers); without stripping those, two halves of one ruling
# never get grouped together because their `_cm_norm` values differ.
_PART_SUFFIX_RES = [
    # "(Part 2 of 2)" / "(Pt. 1/3)" — numeric, may carry trailing cruft
    # like "(Part 2 of 2 for purposes of entry of tentative ruling only)".
    # Stop at the closing paren OR end-of-string so the broader admin
    # pattern below doesn't have to relitigate this.
    re.compile(r'\s*\(?\s*(?:part|pt)\.?\s+\d+\s+(?:of|/)\s+\d+\b[^)]*\)?\s*\.?\s*$',
               re.IGNORECASE),
    # "(Part two of two)" / "(part one of two)" — spelled-out numbers
    # (one through six covers everything we've seen in the corpus).
    re.compile(r'\s*\(?\s*(?:part|pt)\.?\s+(?:one|two|three|four|five|six)\s+'
               r'(?:of|/)\s+(?:one|two|three|four|five|six)\b[^)]*\)?\s*\.?\s*$',
               re.IGNORECASE),
    # ANY trailing parenthetical that mentions "tentative ruling" — the
    # SFTC backend emits a long tail of admin annotations on the second
    # half's calendar_matter, all of which boil down to "this row is the
    # entry/posting half of a multi-part ruling". Variants seen in the
    # wild include:
    #   (ADDED TO CALENDAR FOR TENTATIVE RULING ENTRY PURPOSES ONLY)
    #   (Added For Posting Of Tentative Ruling)
    #   (Added For Posting Of Tentative Ruling Only)
    #   (Added To Calendar For Issuance Of Tentative Ruling Only)
    #   (Added To Calendar For Issuance Of Tentative Ruling Purposes Only)
    #   (Added To Calendar For Tentative Ruling)
    #   (Added For Posting Of Tentative Ruling.)
    #   (For Purposes Of Entry Of Tentative Ruling)
    # Matching on the literal phrase "tentative ruling" inside parens
    # catches every variant without requiring per-form regex maintenance.
    re.compile(r'\s*\([^)]*\btentative\s+ruling\b[^)]*\)\s*\.?\s*$',
               re.IGNORECASE),
    # Same idea without surrounding parens (rarer but does happen):
    #   "...; for purposes of entry of tentative ruling"
    #   "... added to calendar for tentative ruling entry purposes only"
    re.compile(r'\s*[;,]?\s*(?:added\s+(?:to\s+calendar|for\s+posting|for\s+issuance)|for\s+purposes\s+of\s+(?:entry|posting))[^()]*\btentative\s+ruling\b[^()]*$',
               re.IGNORECASE),
    # "(Added For Posting)" / "(Added To Calendar For Posting)" —
    # admin annotations that omit the "tentative ruling" phrase but
    # carry the same meaning. Must mention "added" + "posting/issuance/entry"
    # to avoid overshadowing legitimate motion titles like "Added Defendant".
    re.compile(r'\s*\(\s*added\s+(?:to\s+calendar\s+)?(?:for\s+)?(?:posting|issuance|entry)[^)]*\)\s*\.?\s*$',
               re.IGNORECASE),
    # "(continued from previous entry)" / "(see previous entry)" /
    # "(continues, see next entry)" / "(tentative ruling continues …)"
    # — narrative continuation markers without explicit part numbering.
    re.compile(r'\s*\(?\s*(?:tentative\s+ruling\s+)?continu(?:es|ed)\b[^)]*\)?\s*\.?\s*$',
               re.IGNORECASE),
    re.compile(r'\s*\(?\s*see\s+(?:next|previous|prior)\s+entry\s*\)?\s*\.?\s*$',
               re.IGNORECASE),
]
_PART_SUFFIX_RE = _PART_SUFFIX_RES[0]  # back-compat: the numeric form is
                                       # what other call sites rewrite into.

# Numeric and word-form part-number extractors. Used to sort the parts
# of a split ruling into the right order before concatenation. Matches
# anywhere in the text (the marker may live in the calendar_matter
# suffix or inside the ruling body).
_PART_NUM_RE = re.compile(
    r'\(?\s*(?:part|pt)\.?\s+(\d+)\s+(?:of|/)\s+(\d+)\s*\)?',
    re.IGNORECASE,
)
_PART_WORD_RE = re.compile(
    r'\(?\s*(?:part|pt)\.?\s+(one|two|three|four|five|six)\s+(?:of|/)\s+(one|two|three|four|five|six)\s*\)?',
    re.IGNORECASE,
)
_WORD_TO_NUM = {'one': 1, 'two': 2, 'three': 3, 'four': 4, 'five': 5, 'six': 6}


def extract_part_total(text):
    """Total number of parts (the M in "Part N of M"). Used by the
    consolidator to decide whether all parts have arrived for a given
    case+date — a strong signal that the rows belong together even when
    their captions diverge."""
    if not text:
        return None
    m = _PART_NUM_RE.search(text)
    if m:
        return int(m.group(2))
    m = _PART_WORD_RE.search(text)
    if m:
        return _WORD_TO_NUM.get(m.group(2).lower())
    return None


def normalize_motion_for_split(cm):
    """Canonicalise a calendar_matter for split-ruling grouping.

    Strips every recognised continuation suffix (numeric Part N/M,
    spelled-out Part two of two, the "added to calendar for tentative
    ruling entry purposes only" tag, and the bare "continues..." /
    "see next entry" narrative markers). Two halves of the same ruling
    typically share most of their calendar_matter once these suffixes
    are gone, which is what `consolidate_splits` keys on.
    """
    if not cm:
        return ""
    # Apply each suffix-stripping regex in turn, possibly twice if the
    # tail has stacked tags (e.g. "Foo (Part 2 of 2). (continues…)").
    out = cm
    for _ in range(2):
        for r in _PART_SUFFIX_RES:
            out = r.sub("", out)
    return out.strip().lower()


def extract_part_num(text):
    if not text:
        return None
    m = _PART_NUM_RE.search(text)
    if m:
        return int(m.group(1))
    m = _PART_WORD_RE.search(text)
    if m:
        return _WORD_TO_NUM.get(m.group(1).lower())
    return None


def make_hash(case_number, court_date, ruling):
    key = f"{case_number}|{court_date}|{ruling or ''}"
    return hashlib.sha1(key.encode()).hexdigest()


def normalize_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d %I:%M %p", "%b-%d-%Y %I:%M %p", "%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(val.strip(), fmt).date().isoformat()
            except ValueError:
                continue
        # e.g. "JAN-02-2020 09:00 AM" — try just the date part
        try:
            return datetime.strptime(val.strip().split()[0], "%b-%d-%Y").date().isoformat()
        except ValueError:
            pass
        # e.g. "2026-04-27 09:00 AM" — try just the date part
        try:
            return datetime.strptime(val.strip().split()[0], "%Y-%m-%d").date().isoformat()
        except ValueError:
            return val
    return str(val)


def normalize_time(val):
    if val is None:
        return None
    if isinstance(val, time):
        return val.strftime("%H:%M")
    if isinstance(val, datetime):
        return val.strftime("%H:%M")
    if isinstance(val, str):
        # "09:00 AM" or "2026-04-27 09:00 AM"
        parts = val.strip().split()
        if len(parts) >= 3:
            try:
                return datetime.strptime(f"{parts[1]} {parts[2]}", "%I:%M %p").strftime("%H:%M")
            except ValueError:
                pass
        if len(parts) == 2:
            try:
                return datetime.strptime(f"{parts[0]} {parts[1]}", "%I:%M %p").strftime("%H:%M")
            except ValueError:
                pass
    return str(val) if val else None


def load_xlsm_2014_2018(path, department="302"):
    wb = openpyxl.load_workbook(path, read_only=True, keep_vba=True)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        case_number, case_title, court_date, calendar_matter, ruling = r[:5]
        if not any([case_number, ruling]):
            continue
        hearing_time = None
        if isinstance(court_date, datetime) and (court_date.hour or court_date.minute):
            hearing_time = court_date.strftime("%H:%M")
        rows.append({
            "department":      department,
            "case_number":     str(case_number).strip() if case_number else None,
            "case_title":      str(case_title).strip() if case_title else None,
            "court_date":      normalize_date(court_date),
            "hearing_time":    hearing_time,
            "calendar_matter": str(calendar_matter).strip() if calendar_matter else None,
            "judge":           None,
            "ruling":          str(ruling).strip() if ruling else None,
        })
    return rows


def load_xlsx_2020_plus(path, department="302"):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["tentatives"]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        case_number, case_title, court_date, hearing_time, calendar_matter, judge, ruling = r[:7]
        if not any([case_number, ruling]):
            continue
        rows.append({
            "department":      department,
            "case_number":     str(case_number).strip() if case_number else None,
            "case_title":      str(case_title).strip() if case_title else None,
            "court_date":      normalize_date(court_date),
            "hearing_time":    normalize_time(hearing_time),
            "calendar_matter": str(calendar_matter).strip() if calendar_matter else None,
            "judge":           normalize_judge_name(str(judge).strip()) if judge else None,
            "ruling":          str(ruling).strip() if ruling else None,
        })
    return rows


def _calendar_kind_from_path(path):
    """Infer the Asbestos sub-calendar from a raw file's path. Dept 304
    files live under raw/dept304/<kind>/ where <kind> is 'discovery' or
    'law-and-motion'; legacy flat-layout files in raw/dept304/ predate
    the split and are all asbestos-law-and-motion (verified by ruling
    text). Other depts have no sub-calendars and return None."""
    parts = Path(path).parts
    for i, p in enumerate(parts):
        if p.startswith("dept"):
            after = parts[i + 1] if i + 1 < len(parts) else ""
            if after and not after.endswith(".json"):
                return after  # 'discovery' or 'law-and-motion'
            if p == "dept304":
                return "law-and-motion"  # legacy flat-layout fallback
            break
    return None


def load_json(path, department=None):
    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    # Support both formats:
    #   Legacy: [{...}, ...]                                — uses `department`
    #   Extension: {department, rulings: [{...}, ...]}      — wrapper wins
    if isinstance(data, list):
        records = data
        if department is None:
            department = "302"
        wrapper_kind = None
    else:
        records    = data.get("rulings", [])
        # Wrapper department takes precedence; CLI --dept is a fallback for
        # malformed scrapes that omit the field.
        department = str(data.get("department") or department or "302")
        wrapper_kind = data.get("calendar_kind")

    # Sub-calendar tag: prefer the wrapper's explicit field (newer
    # extension scrapes set it), fall back to inferring from the raw
    # file's path so the legacy flat-layout dept 304 files still get
    # tagged correctly without a one-time backfill pass.
    calendar_kind = wrapper_kind or _calendar_kind_from_path(path)

    rows = []
    for rec in records:
        court_date_raw = rec.get("Court Date", "")
        ruling_text    = rec.get("Rulings", "").strip() or None
        # Use explicit Judge field if present (scraped by extension); for Probate
        # records the extension folds the Examiner field into Judge (title-cased),
        # but fall back to Examiner directly in case an older scrape kept it raw.
        examiner_raw = (rec.get("Examiner") or "").strip()
        judge = rec.get("Judge") or (examiner_raw.title() if examiner_raw else None) or extract_judge(ruling_text)
        rows.append({
            "department":      department,
            "case_number":     rec.get("Case Number", "").strip() or None,
            "case_title":      rec.get("Case Title", "").strip() or None,
            "court_date":      normalize_date(court_date_raw),
            "hearing_time":    normalize_time(court_date_raw),
            "calendar_matter": rec.get("Calendar Matter", "").strip() or None,
            "judge":           normalize_judge_name(judge),
            "ruling":          ruling_text,
            "calendar_kind":   calendar_kind,
        })
    return rows


def detect_and_load(path, department="302"):
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".json":
        return load_json(path, department=department)
    if suffix in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(path, read_only=True, keep_vba=(suffix == ".xlsm"))
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        wb.close()
        if "Judge" in headers or "Hearing Time" in headers:
            return load_xlsx_2020_plus(path, department=department)
        return load_xlsm_2014_2018(path, department=department)
    raise ValueError(f"Unsupported file type: {path}")


def to_df(rows):
    df = pd.DataFrame(rows, columns=[c for c in COLUMNS if c != "row_hash"])
    return _add_hash(df)


def _add_hash(df):
    df["row_hash"] = df.apply(
        lambda r: make_hash(r["case_number"] or "", r["court_date"] or "", r["ruling"] or ""),
        axis=1,
    )
    return df[COLUMNS]


_DEDUPE_NORM_RE = re.compile(r'[^\w\s]')

def _normalize_for_dedupe(s):
    """Strip whitespace + punctuation variation so two near-identical
    paragraphs (one with a stray period, one without) collapse into a
    single canonical form for substring comparison. SFTC sometimes
    re-emits a Part 1 inside a Part 2 with minor formatting drift —
    the byte-exact dedupe missed those, leaving the user with two
    near-copies of the same ruling text concatenated."""
    if not s:
        return ""
    s = _WS_RE.sub(' ', s.lower())
    s = _DEDUPE_NORM_RE.sub('', s)
    return s.strip()


def dedupe_paragraphs(text):
    """Drop paragraphs that are near-duplicates of an earlier paragraph
    in the same ruling. SFTC sometimes emits a Part 1 whose text already
    contains the substantive ruling, then a Part 2 that repeats the same
    text with a minor formatting tweak (a trailing period dropped, a
    space collapsed). Past consolidations concatenated both, leaving
    rulings whose two halves are byte-different but content-identical.
    The split-time `_dedupe_near_duplicates` catches this between
    separate rulings, but a row already merged in the parquet is past
    that gate — so we run a paragraph-level pass on every ruling text
    too. Idempotent: a row with no internal duplication comes back
    unchanged."""
    if not text:
        return text
    paras = text.split("\n\n")
    if len(paras) < 2:
        return text
    seen_norms = []
    kept_paras = []
    for p in paras:
        if not p.strip():
            continue
        n = _normalize_for_dedupe(p)
        if not n:
            continue
        # Skip if this paragraph's normalised form is contained in any
        # already-kept paragraph (or vice versa). Symmetric containment
        # — a longer late paragraph that contains an earlier one drops
        # the earlier one too. Two-pass: once forward to flag, once
        # backward to compose the result.
        if any(n in s or s in n for s in seen_norms):
            continue
        seen_norms.append(n)
        kept_paras.append(p)
    return "\n\n".join(kept_paras) if kept_paras else text


def _dedupe_near_duplicates(seq):
    """Drop strings that are near-duplicates (after whitespace and
    punctuation normalisation) of any longer string in `seq`. Replaces
    the byte-exact containment check, which left behind paragraph
    pairs that differed only in a trailing period — see bug #9."""
    candidates = sorted({s for s in seq if s and s.strip()}, key=len, reverse=True)
    kept = []
    kept_norms = []
    for s in candidates:
        n = _normalize_for_dedupe(s)
        if not n:
            continue
        # Skip if its normalised form is contained in any already-kept
        # entry's normalised form. The longer entry wins.
        if any(n in kn for kn in kept_norms):
            continue
        kept.append(s)
        kept_norms.append(n)
    return kept


def consolidate_splits(df: pd.DataFrame) -> pd.DataFrame:
    """Concatenate split rulings.

    The SFTC site sometimes truncates a long ruling and emits the rest as a
    second record marked "(Part 2 of 2)" (or with the suffix on
    calendar_matter). Two rows are parts of the same ruling when they
    share (case_number, court_date) AND either:

      • their calendar_matter is identical after part-suffix stripping
        (the canonical case — both halves share the original caption); or
      • at least one row in the (case_number, court_date) group carries
        an explicit "Part N of M" marker (in either the ruling text or
        the calendar_matter). The marker is unambiguous: there should
        be exactly M parts on this case+date for this motion, so we
        merge the whole (case, date) group regardless of caption
        variation. This catches the common SFTC pattern where the
        Part 2 row's caption is "DEFENDANT X, Y, Z, ... <motion>. Part
        2 of 2" while Part 1's caption is just "<motion>" — the part
        suffix stripping isn't enough to align them, but the explicit
        "(Part N of M)" marker is enough on its own.

    Within each group we sort by part number (taken from ruling text
    or calendar_matter) and join with two newlines, with whitespace +
    punctuation-aware dedupe to stop near-duplicate paragraphs from
    landing in the merged output.
    """
    if df.empty:
        return df
    df = df.copy()
    df["_cm_norm"] = df["calendar_matter"].fillna("").apply(normalize_motion_for_split)
    df["_part"]  = df["ruling"].fillna("").apply(extract_part_num)
    df["_part"]  = df["_part"].fillna(df["calendar_matter"].fillna("").apply(extract_part_num))
    df["_total"] = df["ruling"].fillna("").apply(extract_part_total)
    df["_total"] = df["_total"].fillna(df["calendar_matter"].fillna("").apply(extract_part_total))

    # Stable order before grouping so single-part rows keep their original placement.
    df = df.sort_values(["case_number", "court_date", "_cm_norm", "_part"],
                        na_position="last", kind="stable").reset_index(drop=True)

    # (case_number, court_date) groups whose rows carry an explicit
    # "Part N of M" marker get the broader (case, date) grouping. We
    # implement the override by zeroing _cm_norm on every such row —
    # then the existing 3-key groupby naturally collapses the whole
    # (case, date) bucket into one group, and pandas does the heavy
    # lifting in C without a Python-level per-row loop.
    pair_part_any = (df.assign(_has_part=df["_part"].notna())
                       .groupby(["case_number", "court_date"], dropna=False)["_has_part"]
                       .transform("any"))
    df.loc[pair_part_any.fillna(False), "_cm_norm"] = ""

    keys = ["case_number", "court_date", "_cm_norm"]
    sizes = df.groupby(keys, dropna=False).size()
    multi_keys = set(sizes[sizes > 1].index)

    if not multi_keys:
        return df.drop(columns=["_cm_norm", "_part", "_total"], errors="ignore")

    groups_map = df.groupby(keys, dropna=False).groups

    # Single-row groups stay as-is; multi-row groups get consolidated.
    keep_indices = []
    merged_rows = []
    consolidated = 0
    for key, idx in groups_map.items():
        if len(idx) <= 1:
            keep_indices.extend(idx)
            continue
        sub = df.loc[idx]
        raw_rulings = [r for r in sub["ruling"].fillna("").tolist() if r.strip()]
        rulings = _dedupe_near_duplicates(raw_rulings)
        if len(rulings) <= 1:
            # Already-merged or all-near-duplicate case — keep one row
            # whose ruling is the longest (the post-consolidation form,
            # if present).
            longest_idx = max(idx, key=lambda i:
                              len(df.at[i, "ruling"]) if isinstance(df.at[i, "ruling"], str) else 0)
            keep_indices.append(longest_idx)
            consolidated += len(idx) - 1
            continue
        merged_text = dedupe_paragraphs("\n\n".join(rulings))
        # When the group was merged via the part-marker override, the
        # captions can differ enough that picking sub.iloc[0] would
        # leave a "Part 2 of 2" caption on the merged row. Prefer the
        # SHORTEST caption (after part-suffix stripping) — that's
        # almost always the bare motion title without the defendant
        # list / annotations.
        candidate_idx = sorted(idx, key=lambda i: (
            len(_PART_SUFFIX_RE.sub("", df.at[i, "calendar_matter"] or "").strip())
                if isinstance(df.at[i, "calendar_matter"], str) else 999_999,
            i,
        ))
        first = df.loc[candidate_idx[0]].to_dict()
        first["ruling"] = merged_text
        cm = first.get("calendar_matter")
        cm = cm if isinstance(cm, str) else ""
        first["calendar_matter"] = normalize_motion_for_split(cm).strip().title() or None
        # `normalize_motion_for_split` lowercases for comparison; rebuild a
        # display-cased form so the merged row's caption isn't all-lowercase.
        # Falls back to the raw stripped form if title-casing destroys
        # acronyms (rare — SFTC captions are mostly title-case already).
        if isinstance(first["calendar_matter"], str) and first["calendar_matter"]:
            display_cm = _PART_SUFFIX_RE.sub("", df.at[candidate_idx[0], "calendar_matter"] or "").strip()
            first["calendar_matter"] = display_cm or None
        merged_rows.append(first)
        consolidated += len(idx) - 1

    base = df.loc[keep_indices].drop(columns=["_cm_norm", "_part", "_total"], errors="ignore")
    if merged_rows:
        merged_df = pd.DataFrame(merged_rows).drop(
            columns=["_cm_norm", "_part", "_total"], errors="ignore")
        out = pd.concat([base, merged_df], ignore_index=True)
    else:
        out = base
    out = _add_hash(out.drop(columns=[c for c in out.columns if c == "row_hash"], errors="ignore"))
    out = out.drop_duplicates(subset="row_hash", keep="first").reset_index(drop=True)
    if consolidated:
        print(f"  consolidated {consolidated} split-ruling row(s) into "
              f"{len(merged_rows)} merged ruling(s)")
    return out


def migrate_existing(df: pd.DataFrame) -> pd.DataFrame:
    """Bring an existing parquet up to the current schema:

      - Drop the legacy `admin_notes` column, folding its text back into
        `ruling` so the full ruling is in one place (browser splits at
        display time now).
      - Recompute `row_hash` from the full ruling.
      - Concatenate split rulings ((Part N of M)) into one record.
    """
    if "department" not in df.columns:
        # One-time migration: parquets predating the multi-dept schema only
        # held Civil Law and Motion (Dept 302) data. New ingests always carry
        # department through, so this branch is never hit twice.
        df.insert(0, "department", "302")

    if "admin_notes" in df.columns:
        # Reattach admin text to the ruling. Old splits always had the boundary
        # phrase at the start of admin_notes, so a single-space join recovers
        # the original substring.
        def rejoin(row):
            r = row.get("ruling")
            a = row.get("admin_notes")
            r = r if isinstance(r, str) else ""
            a = a if isinstance(a, str) else ""
            if r and a:
                return f"{r.rstrip()} {a.lstrip()}"
            return r or a or None
        df["ruling"] = df.apply(rejoin, axis=1)
        df = df.drop(columns=["admin_notes"])

    # Normalise judge names so the Excel "Richard B. Ulmer, Jr." and the
    # extension-scrape "Richard B. Ulmer Jr." stop appearing as two rows in
    # the dropdown / charts.
    if "judge" in df.columns:
        df["judge"] = df["judge"].apply(normalize_judge_name)

    # Paragraph-level dedup pass over every existing ruling. Catches the
    # back-catalogue case where a previous merge concatenated two
    # near-identical paragraphs into one row (the byte-exact dedupe of
    # the day didn't see them as substrings of each other). Idempotent
    # on rows without intra-string duplication, so safe to run on every
    # --all-raw pass.
    if "ruling" in df.columns:
        df["ruling"] = df["ruling"].apply(dedupe_paragraphs)

    # Backfill any missing columns so the column-order projection at the
    # bottom doesn't blow up on a parquet that predates a recently-added
    # schema column.
    for col in COLUMNS:
        if col not in df.columns:
            df[col] = None

    # Calendar-kind backfill for Dept 304 rows imported before the column
    # existed. Discovery sub-calendar rulings always begin "On Asbestos
    # Discovery Calendar..."; everything else in 304 is asbestos-law-and-
    # motion. Path-based inference handles future ingests; this text-based
    # pass closes the gap on already-indexed rows.
    if "department" in df.columns and "calendar_kind" in df.columns:
        m304 = (df["department"] == "304") & df["calendar_kind"].isna()
        if m304.any():
            disc_re = re.compile(r"On\s+Asbestos\s+Discovery\s+Calendar", re.IGNORECASE)
            df.loc[m304 & df["ruling"].fillna("").str.contains(disc_re),
                   "calendar_kind"] = "discovery"
            df.loc[(df["department"] == "304") & df["calendar_kind"].isna(),
                   "calendar_kind"] = "law-and-motion"

    # consolidate_splits already recomputes the hash and dedupes; we just need
    # to ensure the canonical column order/presence on its way out.
    df = consolidate_splits(df)
    return df[COLUMNS].reset_index(drop=True)


def merge(existing: pd.DataFrame, new: pd.DataFrame):
    combined = pd.concat([existing, new], ignore_index=True)
    combined = combined.drop_duplicates(subset="row_hash", keep="first")
    # Re-run consolidation after merge — a newly-arrived part may pair with an
    # existing row in the parquet.
    combined = consolidate_splits(combined)
    after    = len(combined)
    inserted = after - len(existing)
    skipped  = len(new) - inserted
    return combined.sort_values(["court_date", "department"]).reset_index(drop=True), inserted, skipped


def add_split_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Run split_ruling against each row's ruling text and store the
    three result strings as separate columns: ruling_substantive,
    ruling_admin, ruling_courtcall. The original `ruling` column is
    preserved untouched for back-compat with anyone scripting against
    the canonical parquet directly. update-readme.py uses these
    pre-computed columns to slice the per-dept main parquet
    (substantive only) from the per-dept extras parquet (admin +
    courtcall, lazy-loaded by the data browser)."""
    if df.empty:
        df["ruling_substantive"] = ""
        df["ruling_admin"]       = ""
        df["ruling_courtcall"]   = ""
        return df
    splits = df["ruling"].fillna("").apply(split_ruling)
    df = df.copy()
    df["ruling_substantive"] = splits.apply(lambda t: t[0])
    df["ruling_admin"]       = splits.apply(lambda t: t[1])
    df["ruling_courtcall"]   = splits.apply(lambda t: t[2])
    return df


def save_parquet(df: pd.DataFrame):
    df.to_parquet(PARQUET, index=False, compression="zstd")


def save_sqlite(df: pd.DataFrame):
    conn = sqlite3.connect(DB_PATH)
    conn.executescript("""
        DROP TABLE IF EXISTS tentatives;
        CREATE TABLE tentatives (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            department      TEXT,
            case_number     TEXT,
            case_title      TEXT,
            court_date      DATE,
            hearing_time    TEXT,
            calendar_matter TEXT,
            judge           TEXT,
            ruling          TEXT,
            row_hash        TEXT UNIQUE,
            calendar_kind   TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_department    ON tentatives(department);
        CREATE INDEX IF NOT EXISTS idx_case_number   ON tentatives(case_number);
        CREATE INDEX IF NOT EXISTS idx_court_date    ON tentatives(court_date);
        CREATE INDEX IF NOT EXISTS idx_judge         ON tentatives(judge);
    """)
    df.to_sql("tentatives", conn, if_exists="append", index=False,
              method="multi", chunksize=500)
    conn.close()


def summary(df: pd.DataFrame):
    print(f"\nDataset: {len(df):,} rows  |  {df['court_date'].min()} → {df['court_date'].max()}")
    depts = df['department'].dropna().unique()
    print(f"Departments: {', '.join(sorted(depts))}")
    print(f"Parquet: {PARQUET.stat().st_size / 1e6:.1f} MB   DB: {DB_PATH.stat().st_size / 1e6:.1f} MB")


def main():
    parser = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    parser.add_argument("--dept", default="302",
                        help="Department to tag rows with when the source file "
                             "doesn't carry one (Excel imports + legacy bare-list "
                             "JSON). Extension-format JSON wrappers always win.")
    parser.add_argument("--all-raw", action="store_true",
                        help="Ingest every JSON under raw/dept*/. Useful as a "
                             "catch-up when the per-commit GitHub Action falls "
                             "behind a bulk scrape (idempotent — relies on "
                             "row_hash dedup + consolidate_splits).")
    parser.add_argument("paths", nargs="*",
                        help="Source files to ingest. If omitted (and --all-raw "
                             "is not set), falls back to the original Excel "
                             "exports in the repo root.")
    args = parser.parse_args()

    existing = pd.read_parquet(PARQUET) if PARQUET.exists() else pd.DataFrame(columns=COLUMNS)
    existing = migrate_existing(existing)

    if args.all_raw:
        # Walk every dept's raw/*.json. Each file's wrapper carries its own
        # department, so --dept is just a fallback for malformed scrapes.
        # Recursive glob picks up the Dept 304 sub-folders
        # (raw/dept304/law-and-motion/, raw/dept304/discovery/) on top of
        # the legacy flat dept*/*.json layout other departments still use.
        sources = sorted((HERE / "raw").glob("dept*/**/*.json"))
        if not sources:
            print("No raw JSON files found under raw/dept*/")
        else:
            print(f"Catch-up ingest of {len(sources)} raw JSON files…")
    elif args.paths:
        sources = [Path(p) for p in args.paths]
    else:
        sources = [
            HERE / "tentatives.xlsm",
            HERE / "sfsc tentatives 01-2020 to 07-2025.xlsx",
        ]

    if args.all_raw:
        # Bulk path: load every file into one batch and merge once. Calling
        # `merge` per file would re-run `consolidate_splits` over the growing
        # 50K-row frame on every iteration — quadratic in the number of files
        # and unusably slow on a 3000-file backlog.
        print("Loading raw files into a batch…")
        all_rows = []
        for i, p in enumerate(sources, 1):
            if not p.exists():
                continue
            try:
                all_rows.extend(detect_and_load(p, department=args.dept))
            except Exception as e:
                print(f"  ! skipped {p.name}: {e}")
            if i % 250 == 0:
                print(f"  loaded {i}/{len(sources)} files, {len(all_rows)} rows so far")
        new_df = to_df(all_rows)
        print(f"Merging {len(new_df)} new rows into parquet…")
        existing, inserted, skipped = merge(existing, new_df)
        print(f"Total: {inserted} inserted, {skipped} skipped (duplicates)")
    elif len(sources) > 1:
        # Multi-file batch path: same shape as --all-raw but for explicit
        # paths. The per-commit GitHub Action used to feed this loop one
        # file at a time, which made consolidate_splits run quadratically
        # whenever a bulk scrape pushed >10 files at once. One merge keeps
        # the workflow time linear in the file count.
        print(f"Batch ingest of {len(sources)} files…")
        all_rows = []
        for p in sources:
            if not p.exists():
                print(f"  ! not found: {p}")
                continue
            try:
                all_rows.extend(detect_and_load(p, department=args.dept))
            except Exception as e:
                print(f"  ! skipped {p.name}: {e}")
        new_df = to_df(all_rows)
        print(f"Merging {len(new_df)} new rows into parquet…")
        existing, inserted, skipped = merge(existing, new_df)
        print(f"  {inserted} inserted, {skipped} skipped (duplicates)")
    else:
        for p in sources:
            if not p.exists():
                print(f"Not found: {p.name if not args.paths else p}")
                continue
            print(f"Loading {p.name} (dept default = {args.dept})...")
            new_df = to_df(detect_and_load(p, department=args.dept))
            existing, inserted, skipped = merge(existing, new_df)
            print(f"  {inserted} inserted, {skipped} skipped (duplicates)")

    print("Splitting rulings into substantive / admin / courtcall…")
    with_splits = add_split_columns(existing)
    print("Saving parquet...")
    save_parquet(with_splits)
    print("Saving SQLite...")
    # SQLite mirrors the canonical schema only — the split columns live
    # in the parquet for the data browser's slicer. Dropping them here
    # keeps sqlite's CREATE TABLE statement stable.
    save_sqlite(existing)
    summary(existing)


if __name__ == "__main__":
    main()
